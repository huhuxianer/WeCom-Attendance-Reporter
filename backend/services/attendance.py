"""
考勤数据统计报表生成服务

模版结构（1月考勤数据统计_模版.xlsx）：
  Sheet1 列结构：
    Col 1  (A)  : 姓名
    Col 2  (B)  : 身份证号
    Col 3  (C)  : 考勤时间（上午/下午）
    Col 4  (D)  : 星期
    Col 5  (E)  : 1月1日
    ...
    Col 35 (AI) : 1月31日
    Col 36 (AJ) : 备注
    Col 37 (AK) : 出勤（COUNTIF √）
    Col 38 (AL) : 休假（COUNTIF ●）
    Col 39 (AM) : 未打卡（COUNTIF ◎）
    Col 40 (AN) : 事假（COUNTIF ○）
    Col 41 (AO) : 病假（COUNTIF ☆）
    Col 42 (AP) : 迟到（COUNTIF ※）
    Col 43 (AQ) : 早退（COUNTIF ◇）
    Col 44 (AR) : 旷工（COUNTIF ×）
    Col 45 (AS) : 出差（COUNTIF △）
    Col 46 (AT) : 年休（COUNTIF ÷）
    Col 47 (AU) : 调休（COUNTIF ＃）
    Col 48 (AV) : 护理假（COUNTIF ¤）
    Col 49 (AW) : 婚假（COUNTIF Ω）
    Col 50 (AX) : 丧假（COUNTIF ▽）
    Col 51 (AY) : 产假（COUNTIF 👶）

  原模版行号：
    Row 1-3  : 表头区（年月选择、标题、部门）
    Row 4    : 列标题行（姓名/身份证/考勤时间/星期/日期列标题/汇总列标题）
    Row 5    : 日期值行（实际日期，A-C 列与 Row 4 合并，D 列='日'，E-AI 列=日期）
    Row 6-33 : 数据行（14 人，每人 2 行）
    Row 34   : 注释行（合并 A34:AJ34）
    Row 35   : 制表行

  生成逻辑：
    - 删除原 Row 5（日期值行），数据区从 Row 5 开始（原 Row 6 上移）
    - 日期列位置固定：Col 5=1月1日，Col 35=1月31日（day N -> Col 4+N）
    - 每人写 2 行（上午 + 下午），姓名/身份证/备注列合并两行
"""

import copy
import shutil
import calendar
from datetime import datetime
from typing import Optional

import pandas as pd
from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill
from services.config_loader import get_attendance_symbol

# ───────────────────────── 列索引常量（1-based）─────────────────────────
COL_NAME = 1        # A: 姓名
COL_ID = 2          # B: 身份证号
COL_TIME = 3        # C: 考勤时间（上午/下午）
COL_WEEKDAY = 4     # D: 星期
COL_DATE_START = 5  # E: 1月1日（固定：day N -> col = COL_DATE_START + N - 1）
COL_DATE_END = 35   # AI: 1月31日
COL_REMARK = 36     # AJ: 备注

# 汇总列（col_index -> (列名, COUNTIF 符号)）
SUMMARY_COLS = {
    37: ("出勤", "√"),
    38: ("休假", "●"),
    39: ("未打卡", "◎"),
    40: ("事假", "○"),
    41: ("病假", "☆"),
    42: ("迟到", "※"),
    43: ("早退", "◇"),
    44: ("旷工", "×"),
    45: ("出差", "△"),
    46: ("年休", "÷"),
    47: ("调休", "＃"),
    48: ("护理假", "¤"),
    49: ("婚假", "Ω"),
    50: ("丧假", "▽"),
    51: ("产假", "+"),
}

# 颜色常量
WEEKEND_FILL = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
WHITE_FILL = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

# 数据起始行（保留日期行 Row 5 后，数据从 Row 6 开始）
DATA_START_ROW = 6

# 原模版中注释行（Row 34）、制表行（Row 35），删除 Row 5 后各减 1
# 原数据区：Row 6-33 = 28 行（14 人）
# 删除 Row 5 后：数据区 Row 5-32（原 Row 6-33 整体上移 1）
# 注释行原 Row 34 -> 删除后 Row 33
TEMPLATE_DATA_COUNT = 28  # 模版原数据行数（14人 * 2行）


def _parse_day(date_str: str) -> Optional[int]:
    """
    从源数据日期字符串中提取"日"数字。

    格式 "2026/01/15 星期四" -> 15
    返回 None 表示解析失败。
    """
    if not date_str or date_str == "--":
        return None
    try:
        date_part = str(date_str).split()[0]
        dt = datetime.strptime(date_part, "%Y/%m/%d")
        return dt.day
    except Exception:
        return None


def _day_to_col(day: int) -> int:
    """将月中的天数转换为对应的列索引（1-based）"""
    return COL_DATE_START + day - 1


def _make_countif_formula(row: int, symbol: str) -> str:
    """生成 COUNTIF 公式"""
    start = get_column_letter(COL_DATE_START)  # E
    end = get_column_letter(COL_DATE_END)      # AI
    return f'=COUNTIF({start}{row}:{end}{row},"{symbol}")'


def _safe_write(ws, row: int, col: int, value):
    """安全地写入单元格（跳过合并区域的从属单元格）"""
    cell = ws.cell(row, col)
    if not isinstance(cell, MergedCell):
        cell.value = value


def _copy_row_style(ws, src_row: int, dst_row: int):
    """从源行复制样式到目标行（跳过从属合并单元格）"""
    for col in range(1, ws.max_column + 1):
        src = ws.cell(src_row, col)
        dst = ws.cell(dst_row, col)
        if isinstance(dst, MergedCell):
            continue
        if src.has_style:
            dst.font = copy.copy(src.font)
            dst.alignment = copy.copy(src.alignment)
            dst.border = copy.copy(src.border)
            dst.fill = copy.copy(src.fill)
            dst.number_format = src.number_format


def _clear_row(ws, row: int):
    """清空一行所有单元格的值（跳过从属合并单元格）"""
    for col in range(1, ws.max_column + 1):
        cell = ws.cell(row, col)
        if not isinstance(cell, MergedCell):
            cell.value = None


def _remove_merges_in_rows(ws, row_start: int, row_end: int):
    """取消指定行范围内的所有合并单元格"""
    to_remove = []
    for m in ws.merged_cells.ranges:
        # 合并区域与目标范围有重叠则取消
        if m.min_row <= row_end and m.max_row >= row_start:
            to_remove.append(str(m))
    for m_str in to_remove:
        ws.unmerge_cells(m_str)


def _update_header_dates(ws, year: int, month: int):
    """
    动态更新表头中的星期行（Row 4）和日期行（Row 5）。
    """
    days_in_month = calendar.monthrange(year, month)[1]
    weekdays_en = ["Mon", "Tue", "Wed", "Thu", "Fri", "Sat", "Sun"]
    
    for day in range(1, 32):
        col = _day_to_col(day)
        if day <= days_in_month:
            dt = datetime(year, month, day)
            is_we = dt.weekday() >= 5
            ws.cell(4, col).value = weekdays_en[dt.weekday()]
            ws.cell(5, col).value = day
            
            # 设置颜色
            fill = WEEKEND_FILL if is_we else WHITE_FILL
            ws.cell(4, col).fill = fill
            ws.cell(5, col).fill = fill
        else:
            # 清空超出当月天数的列内容
            ws.cell(4, col).value = None
            ws.cell(5, col).value = None
            ws.cell(4, col).fill = WHITE_FILL
            ws.cell(5, col).fill = WHITE_FILL


def _write_summary_formulas(ws, am_row: int, pm_row: int):
    """为上午行和下午行写入汇总 COUNTIF 公式"""
    for col_idx, (name, symbol) in SUMMARY_COLS.items():
        ws.cell(am_row, col_idx).value = _make_countif_formula(am_row, symbol)
        ws.cell(pm_row, col_idx).value = _make_countif_formula(pm_row, symbol)


def _write_person_pair(ws, am_row: int, pm_row: int,
                       name: str, day_am_pm_map: dict,
                       ref_am_row: int, ref_pm_row: int,
                       year: int, month: int):
    """
    写入一人的上午行和下午行数据。

    先复制参考行样式，再写入姓名、考勤时间、各日符号、汇总公式。
    """
    _copy_row_style(ws, ref_am_row, am_row)
    _copy_row_style(ws, ref_pm_row, pm_row)

    # 针对周六周日，强制覆盖颜色
    for day in range(1, 32):
        col_idx = _day_to_col(day)
        try:
            dt = datetime(year, month, day)
            is_we = dt.weekday() >= 5
            fill = WEEKEND_FILL if is_we else WHITE_FILL
            ws.cell(am_row, col_idx).fill = fill
            ws.cell(pm_row, col_idx).fill = fill
        except (ValueError, OverflowError):
            # 超出当月范围，统一涂白
            ws.cell(am_row, col_idx).fill = WHITE_FILL
            ws.cell(pm_row, col_idx).fill = WHITE_FILL

    # 写入姓名、时间标题
    ws.cell(am_row, COL_NAME).value = name
    ws.cell(am_row, COL_TIME).value = "上午"
    ws.cell(pm_row, COL_TIME).value = "下午"

    for day, (am_sym, pm_sym) in day_am_pm_map.items():
        col = _day_to_col(day)
        if COL_DATE_START <= col <= COL_DATE_END:
            ws.cell(am_row, col).value = am_sym
            ws.cell(pm_row, col).value = pm_sym

    _write_summary_formulas(ws, am_row, pm_row)


def generate_attendance_report(
    overview_df: pd.DataFrame,
    template_path: str,
    output_path: str,
    year: int,
    month: int,
    dept: str = None,
    details_df: pd.DataFrame = None,
) -> None:
    """
    根据考勤概况数据生成考勤统计报表，基于模版填充数据。

    参数：
        overview_df   : parse_xlsx 返回的 overview DataFrame
        template_path : 模版 .xlsx 文件路径
        output_path   : 输出 .xlsx 文件路径
        year          : 报表年份
        month         : 报表月份
        dept          : 可选，按部门过滤（包含匹配）
    """
    # ── 1. 复制模版文件 ──
    shutil.copy2(template_path, output_path)

    # ── 2. 过滤数据 ──
    df = overview_df.copy()
    if dept:
        df = df[df["部门"].str.contains(dept, na=False)].copy()

    if df.empty:
        return

    # ── 3. 提取有效日期行，计算每天符号 ──
    df_valid = df[~df["日期"].isin(["", "--", "nan"]) & df["日期"].notna()].copy()
    df_valid["_day"] = df_valid["日期"].apply(_parse_day)
    df_valid = df_valid[df_valid["_day"].notna()].copy()

    # 保持原始顺序的去重姓名列表
    seen: set = set()
    names_ordered = []
    # 重要：这里从全量人员里面查，统计列表中的人员名单应该是输入文件中的所有人员
    for name in df["姓名"]:
        if name not in seen:
            seen.add(name)
            names_ordered.append(name)

    if not names_ordered:
        return

    n_persons = len(names_ordered)
    n_rows_needed = n_persons * 2  #每人 2 行 (针对 Sheet1)

    # ── 4. 加载可写工作簿 ──
    wb = load_workbook(output_path)
    ws = wb.worksheets[0]

    # 修改 Sheet 名字为动态年月
    ws.title = f"{year}.{month}月"
    if len(wb.worksheets) > 1:
        wb.worksheets[1].title = f"{year}.{month}月统计表"

    # ── 5. 更新表头年月 ──
    ws.cell(1, 3).value = year   # C1
    ws.cell(1, 7).value = month  # G1
    _update_header_dates(ws, year, month)
    
    # 刷新 Row 2 的标题
    ws.cell(2, 2).value = f"{year}年{month}月考勤表"
    
    # 刷新 Row 3 的部门名称
    unique_depts = df["部门"].unique()
    if len(unique_depts) == 1:
        dept_display = str(unique_depts[0])
    else:
        dept_display = dept if dept else "全部"
    ws.cell(3, 2).value = dept_display

    # ── 6. 清空数据区 ──
    current_data_end = DATA_START_ROW + TEMPLATE_DATA_COUNT - 1  # = 6 + 28 - 1 = 33
    note_row = current_data_end + 1  # = 34（原注释行）

    # 扩大取消合并的范围，防止新增行进入模版底部合并区域
    _remove_merges_in_rows(ws, DATA_START_ROW, 200) 
    for row in range(DATA_START_ROW, current_data_end + 1):
        _clear_row(ws, row)

    # ── 7. 调整行数（插入或删除以匹配实际人数）──
    if n_rows_needed > TEMPLATE_DATA_COUNT:
        extra = n_rows_needed - TEMPLATE_DATA_COUNT
        ws.insert_rows(note_row, extra)
        note_row += extra
        # 修复插入行的行高（insert_rows 不复制行高，Excel 会显示为极细的行）
        for r in range(note_row - extra, note_row):
            ws.row_dimensions[r].height = 30
    elif n_rows_needed < TEMPLATE_DATA_COUNT:
        extra = TEMPLATE_DATA_COUNT - n_rows_needed
        ws.delete_rows(DATA_START_ROW + n_rows_needed, extra)
        note_row -= extra

    # 参考行（用于复制样式）
    ref_am = DATA_START_ROW
    ref_pm = DATA_START_ROW + 1

    # ── 8. 逐人写入数据 ──
    current_row = DATA_START_ROW
    person_stats = {}
    for name in names_ordered:
        person_ov = df_valid[df_valid["姓名"] == name]
        person_det = details_df[details_df["姓名"] == name] if details_df is not None else pd.DataFrame()

        day_am_pm_map = {}
        # 先用概览数据的状态填充（作为默认或全天状态）
        for _, ov_row in person_ov.iterrows():
            day = ov_row["_day"]
            if day is None: continue
            sym = get_attendance_symbol(str(ov_row.get("考勤结果", "")))
            day_am_pm_map[int(day)] = [sym, sym] # 默认应用到上下午

        # 再用详情数据细化（区分上下午打卡状态）
        if not person_det.empty:
            person_det = person_det.copy()
            # 提取日
            person_det["_day"] = person_det["日期"].apply(_parse_day)
            for day_idx, day_group in person_det.groupby("_day"):
                if day_idx is None: continue
                day_idx = int(day_idx)
                
                # 初始化或获取该日状态
                if day_idx not in day_am_pm_map:
                    day_am_pm_map[day_idx] = ["", ""]
                
                # 分别找上班和下班的最佳匹配
                for _, det_row in day_group.iterrows():
                    punch_type = str(det_row.get("打卡类型", ""))
                    status = str(det_row.get("打卡状态", ""))
                    approval = str(det_row.get("假勤申请", ""))
                    
                    # 优先从“假勤申请”中根据关键词判定符号 (补卡、调休、年假等)
                    leave_sym = None
                    if "补卡" in approval:
                        leave_sym = "√"
                    elif "调休" in approval:
                        leave_sym = "＃"
                    elif "年假" in approval or "年休" in approval:
                        leave_sym = "÷"
                    elif "事假" in approval:
                        leave_sym = "○"
                    elif "病假" in approval:
                        leave_sym = "☆"
                    elif "产假" in approval or "哺乳假" in approval:
                        leave_sym = "+"
                    elif "丧假" in approval:
                        leave_sym = "▽"
                    elif "婚假" in approval:
                        leave_sym = "Ω"
                    elif "出差" in approval or "外出" in approval:
                        leave_sym = "△"
                    elif "护理假" in approval:
                        leave_sym = "¤"
                    
                    # 检查时段是否匹配（例如“下午请假”不应覆盖“上午正常”）
                    matches_period = True
                    if leave_sym:
                        is_am_leave = "上午" in approval
                        is_pm_leave = "下午" in approval
                        # 如果只提到了上午或下午，进行匹配性检查
                        if is_am_leave and not is_pm_leave:
                            if "下班" in punch_type: matches_period = False
                        elif is_pm_leave and not is_am_leave:
                            if "上班" in punch_type: matches_period = False

                    if leave_sym and matches_period:
                        symbol = leave_sym
                    else:
                        # 只有在没有明确匹配的假勤申请时，才跳过无需打卡
                        if "无需" in punch_type:
                            continue
                        symbol = get_attendance_symbol(status)
                    
                    if not symbol:
                        continue

                    # 判断归属（基于打卡类型或时间）
                    if "上班" in punch_type:
                        day_am_pm_map[day_idx][0] = symbol
                    elif "下班" in punch_type:
                        day_am_pm_map[day_idx][1] = symbol
                    elif "外出" in punch_type:
                         # 如果是“外出”，且没有明确时段，默认全天（或根据实际需求调整）
                         # 这里暂且假设外出记录通常是成对出现，或者一条记录代表一次外出
                         # 如果是单条记录且没有上下午标识，可能需要特殊处理。
                         # 但通常“外出”在 details 表里会有时间点，或者“假勤申请”里有时段。
                         # 如果 approval 包含“上午”/“下午”，上面逻辑已经处理了 symbol。
                         # 这里主要处理 punch_type="外出" 且没有被 approval 时段逻辑覆盖的情况。
                         # 简单起见，如果这一行是“外出”，我们就认为它有效。
                         # 但为了不覆盖已有的正确打卡，我们需要小心。
                         # 如果 approval 已经决定了时段（matches_period），那么 symbol 已经有了。
                         # 这里主要是决定填到 [0] 还是 [1]。
                         if "上午" in approval:
                             day_am_pm_map[day_idx][0] = symbol
                         elif "下午" in approval:
                             day_am_pm_map[day_idx][1] = symbol
                         else:
                             # 没有时段，默认全天？或者根据实际打卡时间？
                             # 假设外出单据通常覆盖全天，或者该行数据本身就代表该时段状态
                             day_am_pm_map[day_idx][0] = symbol
                             day_am_pm_map[day_idx][1] = symbol
                    elif "请假" in punch_type:
                        # 如果是“请假”类型行，按时段或双填
                        if "下午" in approval:
                            day_am_pm_map[day_idx][1] = symbol
                        elif "上午" in approval:
                            day_am_pm_map[day_idx][0] = symbol
                        else:
                            day_am_pm_map[day_idx][0] = symbol
                            day_am_pm_map[day_idx][1] = symbol
                    else:
                        # 兜底：如果类型不明确（如只有“无需打卡”），尝试根据时段判定
                        if "下午" in approval:
                            day_am_pm_map[day_idx][1] = symbol
                        elif "上午" in approval:
                            day_am_pm_map[day_idx][0] = symbol
                        else:
                            # 默认如果这行有 symbol 就算是应用到这行
                            # 比如本来没有上班/下班字样只有一行数据
                            day_am_pm_map[day_idx][0] = symbol
                            day_am_pm_map[day_idx][1] = symbol

        am_row = current_row
        pm_row = current_row + 1

        _write_person_pair(
            ws, am_row, pm_row,
            name, day_am_pm_map,
            ref_am, ref_pm,
            year, month
        )

        # ── 重点：从符号中提取统计数据 (用于 Sheet2) ──
        # 合并所有半天符号
        all_syms = []
        for am_pm in day_am_pm_map.values():
            all_syms.extend(am_pm)
        
        def count_syms(target_list):
            return sum(1 for s in all_syms if s in target_list)

        stats = {
            "late": count_syms(["※"]),
            "early": count_syms(["◇"]),
            "trip": count_syms(["△"]) * 0.5,
            "absent": count_syms(["×"]) * 0.5,
            "personal": count_syms(["○"]) * 0.5,
            "sick": count_syms(["☆"]) * 0.5,
            "annual": count_syms(["÷"]) * 0.5,
            "other": count_syms(["Ω", "▽", "+", "¤"]) * 0.5,
            "lieu": count_syms(["＃"]) * 0.5,
            "missing": count_syms(["◎"]) * 0.5,
        }
        person_stats[name] = stats

        # 合并姓名（A）、身份证（B）、备注（AJ）列的上午+下午行
        ws.merge_cells(
            start_row=am_row, start_column=COL_NAME,
            end_row=pm_row, end_column=COL_NAME
        )
        ws.merge_cells(
            start_row=am_row, start_column=COL_ID,
            end_row=pm_row, end_column=COL_ID
        )
        ws.merge_cells(
            start_row=am_row, start_column=COL_REMARK,
            end_row=pm_row, end_column=COL_REMARK
        )

        current_row += 2
        
    # ── 9. 生成 Sheet2 统计表 ──
    if len(wb.worksheets) > 1:
        ws2 = wb.worksheets[1]
        STAT_START_ROW = 2
        
        # 统计表中有数据的行：
        max_row = ws2.max_row
        for r in range(STAT_START_ROW, max_row + 1):
            _clear_row(ws2, r)
            
        current_stat_row = STAT_START_ROW
        for name in names_ordered:
            person_df = df[df["姓名"] == name]
            dept_name = person_df["部门"].iloc[0] if not person_df.empty else ""
            stats = person_stats.get(name, {})

            ws2.cell(current_stat_row, 1).value = dept_name
            ws2.cell(current_stat_row, 2).value = name
            ws2.cell(current_stat_row, 3).value = stats.get("late") or None
            ws2.cell(current_stat_row, 4).value = stats.get("early") or None
            ws2.cell(current_stat_row, 5).value = stats.get("trip") or None
            ws2.cell(current_stat_row, 6).value = stats.get("absent") or None
            ws2.cell(current_stat_row, 7).value = stats.get("personal") or None
            ws2.cell(current_stat_row, 8).value = stats.get("sick") or None
            ws2.cell(current_stat_row, 9).value = stats.get("annual") or None
            ws2.cell(current_stat_row, 10).value = stats.get("other") or None
            ws2.cell(current_stat_row, 11).value = stats.get("lieu") or None
            ws2.cell(current_stat_row, 12).value = stats.get("missing") or None

            # 同步样式
            if current_stat_row > STAT_START_ROW:
                _copy_row_style(ws2, STAT_START_ROW, current_stat_row)

            current_stat_row += 1

    # ── 10. 保存文件 ──
    wb.save(output_path)
    wb.close()
