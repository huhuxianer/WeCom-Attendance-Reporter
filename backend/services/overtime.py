"""
周内加班统计报表生成服务

模版结构（1月周内加班统计_模版.xlsx）：
  Sheet1 列结构：
    Col 1  (A)  : 姓名（每人两行合并）
    Col 2  (B)  : 身份证号（每人两行合并）
    Col 3  (C)  : 考勤时间标识（"20：00-22：00" 或 "22：00之后"）
    Col 4  (D)  : 星期
    Col 5  (E)  : 1月1日
    ...
    Col 35 (AI) : 1月31日
    Col 36 (AJ) : 20:00-22:00 月度次数（COUNTA 公式）
    Col 37 (AK) : 22:00之后 月度次数（COUNTA 公式）

  模版行结构：
    Row 1-3  : 表头区（年月选择、标题、部门）
    Row 4    : 列标题行（姓名/身份证/考勤时间/星期/日期/汇总）
    Row 5    : 日期值行（公式日期，A-C 列与 Row 4 合并）
    Row 6-33 : 数据行（14 人，每人 2 行）
    Row 34   : 制表行

  生成逻辑：
    - 保留模版 Row 1-5 表头结构（不删除日期行）
    - 清空 Row 6-33 数据区（D到AK列）
    - 每人写 2 行：
        第 1 行（奇数）: 时间段 "20：00-22：00"，填写 20:00-22:00 之间下班时间
        第 2 行（偶数）: 时间段 "22：00之后"，填写 22:00 之后下班时间
    - 汇总列写 COUNTA 公式统计非空单元格数（即加班次数）
    - A列和AJ/AK列每人两行合并
"""

import copy
import shutil
import calendar
from datetime import datetime, time
from typing import Optional

import pandas as pd
from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter

# ───────────────────────── 列索引常量（1-based）─────────────────────────
COL_NAME = 1         # A: 姓名
COL_ID = 2           # B: 身份证号
COL_TIME = 3         # C: 考勤时间标识
COL_WEEKDAY = 4      # D: 星期
COL_DATE_START = 5   # E: 1月1日（day N -> col = COL_DATE_START + N - 1）
COL_DATE_END = 35    # AI: 1月31日
COL_SUMMARY_1 = 36   # AJ: 20:00-22:00 月度次数
COL_SUMMARY_2 = 37   # AK: 22:00之后 月度次数

# 时间标识字符串（与模版中完全一致）
LABEL_EARLY = "20：00-22：00"
LABEL_LATE = "22：00之后"

# 颜色常量
WEEKEND_FILL = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
WHITE_FILL = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

# 模版数据区
DATA_START_ROW = 6    # 数据从第 6 行开始
TEMPLATE_DATA_ROWS = 28  # 原模版数据行数（14 人 * 2 行）


# ───────────────────────── 工具函数 ─────────────────────────

def _parse_time(time_str: str) -> Optional[time]:
    """解析实际打卡时间字符串为 time 对象。"""
    if not time_str or str(time_str).strip() in ("--", "nan", ""):
        return None
    t = str(time_str).strip()
    for fmt in ["%H:%M", "%H:%M:%S"]:
        try:
            return datetime.strptime(t, fmt).time()
        except Exception:
            continue
    return None


def _parse_day(date_str: str) -> Optional[int]:
    """从日期字符串中提取"日"数字。格式 "2026/01/15 星期四" -> 15"""
    if not date_str or str(date_str).strip() in ("--", "nan", ""):
        return None
    part = str(date_str).split()[0]
    for fmt in ["%Y/%m/%d", "%Y-%m-%d"]:
        try:
            return datetime.strptime(part, fmt).day
        except Exception:
            continue
    return None


def _day_to_col(day: int) -> int:
    """将月中的天数转换为对应的列索引（1-based）"""
    return COL_DATE_START + day - 1


def _time_to_label(t: time) -> Optional[str]:
    """
    根据打卡时间判断所属时间段标签。
    返回 LABEL_EARLY (20:00-22:00) 或 LABEL_LATE (22:00之后) 或 None（未加班）。
    """
    if t is None:
        return None
    if t >= time(20, 0) and t < time(22, 0):
        return LABEL_EARLY
    if t >= time(22, 0):
        return LABEL_LATE
    return None


def _safe_write(ws, row: int, col: int, value):
    """安全地写入单元格（跳过合并区域的从属单元格）"""
    cell = ws.cell(row, col)
    if not isinstance(cell, MergedCell):
        cell.value = value


def _copy_row_style(ws, src_row: int, dst_row: int):
    """从源行复制样式到目标行（跳过从属合并单元格）"""
    for col in range(1, ws.max_column + 1):
        src = ws.cell(src_row, col)
        dst = ws.cell(dst_row, col)
        if isinstance(dst, MergedCell):
            continue
        if src.has_style:
            dst.font = copy.copy(src.font)
            dst.alignment = copy.copy(src.alignment)
            dst.border = copy.copy(src.border)
            dst.fill = copy.copy(src.fill)
            dst.number_format = src.number_format


def _clear_row_range(ws, row: int, col_start: int, col_end: int):
    """清空指定行指定列范围内单元格的值（跳过从属合并单元格）"""
    for col in range(col_start, col_end + 1):
        cell = ws.cell(row, col)
        if not isinstance(cell, MergedCell):
            cell.value = None


def _remove_merges_in_rows(ws, row_start: int, row_end: int):
    """取消指定行范围内的所有合并单元格"""
    to_remove = []
    for m in ws.merged_cells.ranges:
        if m.min_row <= row_end and m.max_row >= row_start:
            to_remove.append(str(m))
    for m_str in to_remove:
        ws.unmerge_cells(m_str)


def _update_header_dates(ws, year: int, month: int):
    """
    动态更新表头中的星期行（Row 4）和日期行（Row 5）。
    """
    days_in_month = calendar.monthrange(year, month)[1]
    weekdays_en = ["Mon", "Tue", "Wed", "Thu", "Fri", "Sat", "Sun"]
    
    for day in range(1, 32):
        col = _day_to_col(day)
        if day <= days_in_month:
            dt = datetime(year, month, day)
            is_we = dt.weekday() >= 5
            ws.cell(4, col).value = weekdays_en[dt.weekday()]
            ws.cell(5, col).value = day
            
            # 设置颜色
            fill = WEEKEND_FILL if is_we else WHITE_FILL
            ws.cell(4, col).fill = fill
            ws.cell(5, col).fill = fill
        else:
            # 清空超出当月天数的列内容
            ws.cell(4, col).value = None
            ws.cell(5, col).value = None
            ws.cell(4, col).fill = WHITE_FILL
            ws.cell(5, col).fill = WHITE_FILL


def _make_counta_formula(row: int, col_start: int, col_end: int) -> str:
    """生成 COUNTA 公式，统计指定行日期范围内的非空单元格数"""
    start_letter = get_column_letter(col_start)
    end_letter = get_column_letter(col_end)
    return f"=COUNTA({start_letter}{row}:{end_letter}{row})"


# ───────────────────────── 主函数 ─────────────────────────

def generate_overtime_report(
    details_df: pd.DataFrame,
    template_path: str,
    output_path: str,
    year: int,
    month: int,
    dept: str = None,
    overview_df: pd.DataFrame = None,
) -> None:
    """
    根据打卡详情数据生成周内加班统计报表，基于模版填充数据。

    参数：
        details_df    : parse_xlsx 返回的 details DataFrame
                        字段：日期、姓名、部门、打卡类型、实际打卡时间 等
        template_path : 模版 .xlsx 文件路径
        output_path   : 输出 .xlsx 文件路径
        year          : 报表年份
        month         : 报表月份
        dept          : 可选，按部门过滤（包含匹配）
        overview_df   : 可选，传入 overview DataFrame 用于提取全量人员名单
    """
    # ── 1. 复制模版文件 ──
    shutil.copy2(template_path, output_path)

    # ── 2. 筛选"下班"类型记录 ──
    df_raw = details_df.copy()
    if dept:
        df_raw = df_raw[df_raw["部门"].str.contains(dept, na=False)].copy()

    if df_raw.empty:
        return
        
    df = df_raw[df_raw["打卡类型"].str.contains("下班", na=False)].copy()

    # ── 4. 解析日期（日）和打卡时间，筛选加班记录 ──
    df["_day"] = df["日期"].apply(_parse_day)
    df["_time"] = df["实际打卡时间"].apply(_parse_time)
    df["_label"] = df["_time"].apply(_time_to_label)

    # 只保留有加班的记录（20:00 之后）
    df_overtime = df[df["_label"].notna() & df["_day"].notna()].copy()

    # ── 5. 按照模版中原始姓名顺序排列人员 ──
    # 先读取模版中已有的姓名顺序
    wb_template = load_workbook(template_path)
    ws_template = wb_template.worksheets[0]

    template_names = []
    seen_template = set()
    for row_idx in range(DATA_START_ROW, ws_template.max_row + 1):
        val = ws_template.cell(row_idx, COL_NAME).value
        if val and str(val).strip() not in ("", "nan", "姓名") and val not in seen_template:
            template_names.append(val)
            seen_template.add(val)
    wb_template.close()

    # 从原始数据中获取所有人员，优先使用 overview_df 若提供
    all_names = set()
    person_dept_map = {}
    if overview_df is not None:
        ov_df = overview_df.copy()
        if dept:
            ov_df = ov_df[ov_df["部门"].str.contains(dept, na=False)].copy()
        for idx, r in ov_df.iterrows():
            n = r["姓名"]
            if pd.notna(n) and str(n).strip() not in ("", "nan"):
                name_str = str(n).strip()
                all_names.add(name_str)
                person_dept_map[name_str] = str(r["部门"]) if pd.notna(r["部门"]) else ""
    else:
        for idx, r in df_raw.iterrows():
            n = r["姓名"]
            if pd.notna(n) and str(n).strip() not in ("", "nan"):
                name_str = str(n).strip()
                all_names.add(name_str)
                person_dept_map[name_str] = str(r["部门"]) if pd.notna(r["部门"]) else ""

    # 按模版顺序排列所有人，再追加模版中没有但出现在原始数据的人
    names_ordered = []
    for name in template_names:
        if name in all_names:
            names_ordered.append(name)
    for name in all_names:
        if name not in names_ordered:
            names_ordered.append(name)

    if not names_ordered:
        # 没有数据，保存空模版
        wb = load_workbook(output_path)
        wb.save(output_path)
        wb.close()
        return

    n_persons = len(names_ordered)
    n_rows_needed = n_persons * 2  # 每人 2 行

    # ── 6. 加载可写工作簿 ──
    wb = load_workbook(output_path)
    ws = wb.worksheets[0]
    
    # 修改 Sheet 名字为动态年月
    ws.title = f"{year}.{month}月"
    if len(wb.worksheets) > 1:
        wb.worksheets[1].title = f"{year}.{month}月统计表"

    # ── 7. 更新表头年月与标题 ──
    ws.cell(1, 3).value = year   # C1
    ws.cell(1, 7).value = month  # G1
    _update_header_dates(ws, year, month)
    
    # 刷新 Row 2 的总量标题（由于 Row 2 写的是公式，通常会自动更新，但为了稳妥重新赋值）
    # 但由于 OpenPyxl 不一定会自动重算公式，如果单元格是 Value 就不行。
    # 根据之前 grep 结果，Row 2 单元格 2 (B2) 是 =C1&E1&G1&I1&"周内加班统计表"
    # 我们直接触发一下重写，或者让它保持公式即可。
    # 之前脚本显示：Cell 2: =C1&E1&G1&I1&"周内加班统计表"
    # 所以只要 C1 和 G1 更新了，公式由于是字符串拼接通常没问题，
    # 但如果不放心可以直接写死 Value：
    ws.cell(2, 2).value = f"{year}年{month}月周内加班统计表"
    
    # 刷新 Row 3 的部门名称
    unique_depts = df["部门"].unique() if not df.empty else []
    if len(unique_depts) == 1:
        dept_display = str(unique_depts[0])
    else:
        dept_display = dept if dept else "全部"
    ws.cell(3, 2).value = dept_display

    # ── 8. 数据统计与写入 ──
    # 取消数据区内原有的合并，然后清空数据区 ──
    current_data_end = DATA_START_ROW + TEMPLATE_DATA_ROWS - 1  # = 33
    # 扩大范围，防止新增行合并干扰
    _remove_merges_in_rows(ws, DATA_START_ROW, 200)
    for row in range(DATA_START_ROW, current_data_end + 1):
        _clear_row_range(ws, row, COL_NAME, COL_SUMMARY_2)

    # 注释/制表行位置（原模版 Row 34）
    note_row = current_data_end + 1  # = 34

    # ── 8. 调整行数（插入或删除以匹配实际人数）──
    if n_rows_needed > TEMPLATE_DATA_ROWS:
        extra = n_rows_needed - TEMPLATE_DATA_ROWS
        ws.insert_rows(note_row, extra)
        note_row += extra
    elif n_rows_needed < TEMPLATE_DATA_ROWS:
        extra = TEMPLATE_DATA_ROWS - n_rows_needed
        ws.delete_rows(DATA_START_ROW + n_rows_needed, extra)
        note_row -= extra

    # 参考行（用于复制样式）
    ref_row_1 = DATA_START_ROW      # "20:00-22:00" 行样式参考
    ref_row_2 = DATA_START_ROW + 1  # "22:00之后" 行样式参考

    # 用于保存每个人两个阶段的统计结果
    overtime_stats = {}

    # ── 9. 逐人写入数据 ──
    current_row = DATA_START_ROW
    for name in names_ordered:
        person_df = df_overtime[df_overtime["姓名"] == name]

        # 分别构建两个时间段的 day -> 实际打卡时间 映射
        early_map: dict = {}   # LABEL_EARLY: 20:00-22:00
        late_map: dict = {}    # LABEL_LATE:  22:00之后

        for _, rec in person_df.iterrows():
            day = int(rec["_day"])
            label = rec["_label"]
            t_obj = rec["_time"]
            # 将 time 对象格式化为 "HH:MM" 字符串
            time_str = t_obj.strftime("%H:%M") if t_obj else None
            if label == LABEL_EARLY:
                early_map[day] = time_str
            elif label == LABEL_LATE:
                late_map[day] = time_str
                
        overtime_stats[name] = {
            "early": len(early_map),
            "late": len(late_map)
        }

        early_row = current_row
        late_row = current_row + 1

        # 复制样式
        _copy_row_style(ws, ref_row_1, early_row)
        _copy_row_style(ws, ref_row_2, late_row)

        # 针对周六周日，强制覆盖颜色
        for day in range(1, 32):
            col_idx = _day_to_col(day)
            try:
                dt = datetime(year, month, day)
                is_we = dt.weekday() >= 5
                fill = WEEKEND_FILL if is_we else WHITE_FILL
            except (ValueError, OverflowError):
                fill = WHITE_FILL
            
            ws.cell(early_row, col_idx).fill = fill
            ws.cell(late_row, col_idx).fill = fill

        # 写入姓名（A列，写在 early_row）
        ws.cell(early_row, COL_NAME).value = name

        # 写入时间段标识（C列）
        ws.cell(early_row, COL_TIME).value = LABEL_EARLY
        ws.cell(late_row, COL_TIME).value = LABEL_LATE

        # 写入各日加班时间
        for day, time_str in early_map.items():
            col = _day_to_col(day)
            if COL_DATE_START <= col <= COL_DATE_END:
                ws.cell(early_row, col).value = time_str

        for day, time_str in late_map.items():
            col = _day_to_col(day)
            if COL_DATE_START <= col <= COL_DATE_END:
                ws.cell(late_row, col).value = time_str

        # 写入汇总 COUNTA 公式（均写到 early_row，因为合并后 late_row 是从属格，值会被丢弃）
        # AJ：统计 early_row（20:00-22:00 时段数据行）的非空单元格数
        ws.cell(early_row, COL_SUMMARY_1).value = _make_counta_formula(
            early_row, COL_DATE_START, COL_DATE_END
        )
        # AK：统计 late_row（22:00之后 时段数据行）的非空单元格数，公式写到主格 early_row
        ws.cell(early_row, COL_SUMMARY_2).value = _make_counta_formula(
            late_row, COL_DATE_START, COL_DATE_END
        )

        # 合并 A列（姓名）每人两行
        ws.merge_cells(
            start_row=early_row, start_column=COL_NAME,
            end_row=late_row, end_column=COL_NAME
        )
        # 合并 B列（身份证号）每人两行
        ws.merge_cells(
            start_row=early_row, start_column=COL_ID,
            end_row=late_row, end_column=COL_ID
        )
        # 合并 AJ列（20:00-22:00 汇总）每人两行
        ws.merge_cells(
            start_row=early_row, start_column=COL_SUMMARY_1,
            end_row=late_row, end_column=COL_SUMMARY_1
        )
        # 合并 AK列（22:00之后 汇总）每人两行
        ws.merge_cells(
            start_row=early_row, start_column=COL_SUMMARY_2,
            end_row=late_row, end_column=COL_SUMMARY_2
        )

        current_row += 2
        
    # ── 10. 生成 Sheet 2 统计表 ──
    if len(wb.worksheets) > 1:
        ws2 = wb.worksheets[1]
        STAT_START_ROW = 2
        
        # 统计表中有数据的行：
        max_row = ws2.max_row
        for r in range(STAT_START_ROW, max_row + 1):
            _clear_row_range(ws2, r, 1, 4)
            
        current_stat_row = STAT_START_ROW
        for name in names_ordered:
            dept_name = person_dept_map.get(name, "")
            
            stats = overtime_stats.get(name, {"early": 0, "late": 0})
            
            ws2.cell(current_stat_row, 1).value = dept_name
            ws2.cell(current_stat_row, 2).value = name
            ws2.cell(current_stat_row, 3).value = stats["early"] if stats["early"] > 0 else None
            ws2.cell(current_stat_row, 4).value = stats["late"] if stats["late"] > 0 else None
            
            if current_stat_row > STAT_START_ROW:
                _copy_row_style(ws2, STAT_START_ROW, current_stat_row)
                
            current_stat_row += 1

    # ── 11. 保存文件 ──
    wb.save(output_path)
    wb.close()
