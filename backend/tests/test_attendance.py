import sys, os
sys.path.insert(0, os.path.dirname(os.path.dirname(__file__)))
from services.parser import parse_xlsx
from services.attendance import generate_attendance_report
import tempfile

SAMPLE_FILE = os.path.join(os.path.dirname(__file__), "../../资料/上下班打卡_日报_20260101-20260131 .xlsx")
TEMPLATE = os.path.join(os.path.dirname(__file__), "../../资料/1月考勤数据统计_模版.xlsx")


def test_generate_creates_file():
    result = parse_xlsx(SAMPLE_FILE)
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
        out_path = tmp.name
    generate_attendance_report(result.overview, TEMPLATE, out_path, 2026, 1)
    assert os.path.exists(out_path)
    assert os.path.getsize(out_path) > 1000
    os.unlink(out_path)


def test_generate_with_dept_filter():
    result = parse_xlsx(SAMPLE_FILE)
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
        out_path = tmp.name
    generate_attendance_report(result.overview, TEMPLATE, out_path, 2026, 1, dept="示例部门")
    assert os.path.exists(out_path)
    assert os.path.getsize(out_path) > 1000
    os.unlink(out_path)


def test_generate_has_person_data():
    """验证生成的报表包含人员数据"""
    result = parse_xlsx(SAMPLE_FILE)
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
        out_path = tmp.name
    generate_attendance_report(result.overview, TEMPLATE, out_path, 2026, 1, dept="示例部门")

    from openpyxl import load_workbook
    wb = load_workbook(out_path)
    ws = wb.worksheets[0]
    name_cell = ws.cell(6, 1).value
    assert name_cell is not None and name_cell != ""
    os.unlink(out_path)
