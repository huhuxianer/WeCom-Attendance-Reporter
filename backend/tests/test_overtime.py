# backend/tests/test_overtime.py
import sys, os
sys.path.insert(0, os.path.dirname(os.path.dirname(__file__)))
from services.parser import parse_xlsx
from services.overtime import generate_overtime_report
import tempfile

SAMPLE_FILE = os.path.join(os.path.dirname(__file__), "../../资料/上下班打卡_日报_20260101-20260131 .xlsx")
TEMPLATE = os.path.join(os.path.dirname(__file__), "../../资料/1月周内加班统计_模版.xlsx")


def test_generate_creates_file():
    result = parse_xlsx(SAMPLE_FILE)
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
        out_path = tmp.name
    generate_overtime_report(result.details, TEMPLATE, out_path, 2026, 1)
    assert os.path.exists(out_path)
    assert os.path.getsize(out_path) > 1000
    os.unlink(out_path)


def test_generate_with_dept_filter():
    result = parse_xlsx(SAMPLE_FILE)
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
        out_path = tmp.name
    generate_overtime_report(result.details, TEMPLATE, out_path, 2026, 1, dept="示例部门")
    assert os.path.exists(out_path)
    assert os.path.getsize(out_path) > 1000
    os.unlink(out_path)


def test_generate_has_person_data():
    """验证生成报表有人员数据"""
    result = parse_xlsx(SAMPLE_FILE)
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
        out_path = tmp.name
    generate_overtime_report(result.details, TEMPLATE, out_path, 2026, 1, dept="示例部门")

    from openpyxl import load_workbook
    wb = load_workbook(out_path)
    ws = wb.worksheets[0]
    # 找到数据起始行，A列应有姓名
    # 遍历找第一个有姓名的行
    found_name = False
    for row_idx in range(4, 15):
        val = ws.cell(row_idx, 1).value
        if val and str(val) not in ("姓名", "NaN", ""):
            found_name = True
            break
    assert found_name, "未找到人员数据"
    os.unlink(out_path)
